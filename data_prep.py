from openpyxl import load_workbook

def parsing_xlsx_file(file: str) -> list:

    wb = load_workbook(f'{file}/{file}.xlsx')
    sheet = wb[wb.sheetnames[0]]

    cleaned_data = []
    for row in sheet.iter_rows(values_only=True, ):
        processed_row = [cell if cell is not None else "" for cell in row]
        cleaned_data.append(processed_row)
    return cleaned_data